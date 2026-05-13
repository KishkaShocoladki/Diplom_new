import json
import os
import textwrap
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd

from backend.extensions import db
from backend.logger_config import get_logger
from backend.models import Experiment, GraphData, Metric

logger = get_logger(__name__)

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_pdf import PdfPages
    import numpy as np
    HAS_MPL = True
except Exception:
    HAS_MPL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


# ── Metadata ─────────────────────────────────────────────────────────────────

# (human_label, explanation, (warn_threshold, good_threshold) or None)
METRIC_META = {
    "accuracy":   ("Точность (Accuracy)",       "Доля верно классифицированных СВ среди всех. 1.0 = идеально, 0.5 = случайное угадывание.", (0.80, 0.90)),
    "auc":        ("AUC-ROC",                    "Площадь под ROC-кривой: способность модели разделять классы. 0.5 = случайно, 1.0 = идеально.", (0.75, 0.90)),
    "f1":         ("F1-Score",                   "Гармоническое среднее Precision и Recall. Лучшая метрика при несбалансированных классах.", (0.70, 0.85)),
    "precision":  ("Precision",                  "Из предсказанных Affected — доля реально Affected. Низкий → много ложных срабатываний.", (0.70, 0.85)),
    "recall":     ("Recall (Sensitivity)",       "Из реальных Affected — доля найденных моделью. Низкий → модель пропускает важные СВ.", (0.70, 0.85)),
    "val_acc":    ("Val Accuracy",               "Точность на валидационной выборке (последняя эпоха).", (0.80, 0.90)),
    "val_auc":    ("Val AUC-ROC",                "AUC-ROC на валидационной выборке.", (0.75, 0.90)),
    "val_f1":     ("Val F1",                     "F1-Score на валидационной выборке.", (0.70, 0.85)),
    "val_precision": ("Val Precision",           "Precision на валидационной выборке.", (0.70, 0.85)),
    "val_recall": ("Val Recall",                 "Recall на валидационной выборке.", (0.70, 0.85)),
    "train_loss": ("Train Loss",                 "Функция потерь на обучающих данных. Монотонное снижение = модель обучается корректно.", None),
    "val_loss":   ("Val Loss",                   "Потери на валидационной выборке. Рост при снижающемся Train Loss → переобучение.", None),
    "total":      ("Всего СВ",                   "Общее число классифицированных структурных вариантов.", None),
    "pred_1":     ("Affected (класс 1)",          "Число СВ, классифицированных как влияющие на экспрессию генов.", None),
    "pred_0":     ("Neutral (класс 0)",           "Число СВ, классифицированных как нейтральные (без влияния на экспрессию).", None),
    "tp":         ("True Positive (TP)",          "Верно обнаруженные Affected СВ. Влияющие варианты, корректно помечены моделью.", None),
    "tn":         ("True Negative (TN)",          "Верно классифицированные Neutral СВ. Нейтральные варианты, не вызвавшие ложной тревоги.", None),
    "fp":         ("False Positive (FP)",         "Neutral СВ, ошибочно принятые за Affected — ложная тревога.", None),
    "fn":         ("False Negative (FN)",         "Affected СВ, пропущенные моделью — предсказаны как нейтральные. Пропуск реального события.", None),
    "train_rows": ("Обучающая выборка",           "Число записей, использованных для обучения.", None),
    "val_rows":   ("Валидационная выборка",        "Число записей в валидационном наборе.", None),
    "test_rows":  ("Тестовая выборка",             "Число записей в тестовом (hold-out) наборе.", None),
}

GRAPH_META = {
    "roc_curve":              ("ROC-кривая",               "FPR vs TPR при разных порогах классификации. AUC — интегральная мера качества. Пунктир — случайный классификатор (AUC = 0.5)."),
    "pr_curve":               ("Precision–Recall кривая",  "Компромисс между точностью и полнотой при разных порогах. Особенно важна при несбалансированных классах."),
    "calibration":            ("Калибровочная кривая",     "Насколько предсказанные вероятности соответствуют реальным частотам событий. Идеал — диагональ."),
    "confusion_matrix":       ("Матрица ошибок",           "Распределение предсказаний по истинным классам: TP, TN, FP, FN."),
    "metrics_vs_threshold":   ("Метрики vs Порог",         "Как меняются Precision, Recall и F1 при изменении порога. Вертикальная линия — стандартный порог 0.5."),
    "confidence_vs_accuracy": ("Уверенность vs Точность",  "Соответствие между уверенностью модели и реальной точностью. Идеал — диагональ."),
}

KIND_LABELS = {
    "train":     "Обучение с нуля",
    "fine_tune": "Дообучение",
    "retrain":   "Переобучение",
    "predict":   "Предсказание",
}
STATUS_LABELS = {
    "completed": "Завершен",
    "failed":    "Ошибка",
    "running":   "Выполняется",
    "pending":   "Ожидание",
}

DARK_BG   = "#1A1A2E"
ACCENT    = "#2B72FB"
GOLD      = "#F2B705"
GREEN     = "#27AE60"
RED       = "#C0392B"
LIGHT_BG  = "#FAFAFA"


# ── PDF page builders ─────────────────────────────────────────────────────────

def _set_style(ax):
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#555555")
    ax.grid(True, alpha=0.25, linestyle="--")


def _title_bar(fig, text, subtitle=""):
    ax = fig.add_axes([0, 0.93, 1, 0.07])
    ax.set_facecolor(DARK_BG)
    ax.axis("off")
    ax.text(0.5, 0.62, text, ha="center", va="center",
            fontsize=14, fontweight="bold", color="black", transform=ax.transAxes)
    if subtitle:
        ax.text(0.5, 0.18, subtitle, ha="center", va="center",
                fontsize=8, color="#000000", transform=ax.transAxes)


def _footer(fig, text):
    ax = fig.add_axes([0, 0, 1, 0.04])
    ax.axis("off")
    ax.set_facecolor("#EEEEEE")
    ax.text(0.5, 0.5, text, ha="center", va="center",
            fontsize=8, color="#999", transform=ax.transAxes)


def _metric_traffic_light(key, value):
    meta = METRIC_META.get(key)
    if not meta or meta[2] is None or not isinstance(value, (int, float)):
        return "#333333"
    lo, hi = meta[2]
    if value >= hi:
        return GREEN
    if value >= lo:
        return GOLD
    return RED


def _draw_cover_page(pdf, exp, dataset_name=""):
    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor(LIGHT_BG)

    # Header
    ax_hdr = fig.add_axes([0, 0.87, 1, 0.13])
    ax_hdr.set_facecolor(DARK_BG)
    ax_hdr.axis("off")
    ax_hdr.text(0.5, 0.65, "ОТЧЕТ ОБ ЭКСПЕРИМЕНТЕ", ha="center", va="center",
                fontsize=22, fontweight="bold", color="black")
    ax_hdr.text(0.5, 0.22, "Модуль классификации структурных вариантов и прогноза экспрессии генов",
                ha="center", va="center", fontsize=10, color="#AAAACC")

    ax = fig.add_axes([0.08, 0.18, 0.84, 0.67])
    ax.axis("off")

    rows = [
        ("Эксперимент",   exp.name or f"#{exp.id}"),
        ("Тип",           KIND_LABELS.get(exp.kind, exp.kind or "—")),
        ("Статус",        STATUS_LABELS.get(exp.status, exp.status or "—")),
        ("Датасет",       dataset_name or (f"ID {exp.dataset_id}" if exp.dataset_id else "—")),
        ("Создан",        exp.created_at.strftime("%d.%m.%Y %H:%M") if exp.created_at else "—"),
        ("Обновлен",      exp.updated_at.strftime("%d.%m.%Y %H:%M") if exp.updated_at else "—"),
        ("Сформирован",   datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]

    y = 0.96
    step = 0.12
    for i, (label, value) in enumerate(rows):
        bg_c = "#F5F5F5" if i % 2 == 0 else "white"
        rect = mpatches.FancyBboxPatch((-0.02, y - step + 0.01), 1.04, step - 0.015,
                                        boxstyle="round,pad=0.008", linewidth=0.4,
                                        edgecolor="#DDDDDD", facecolor=bg_c,
                                        transform=ax.transAxes, clip_on=False)
        ax.add_patch(rect)
        ax.text(0.02, y - step * 0.4, label + ":", fontsize=11, fontweight="bold",
                color="#444", va="center", transform=ax.transAxes)
        ax.text(0.38, y - step * 0.4, str(value), fontsize=11, color="#111",
                va="center", transform=ax.transAxes)
        y -= step

    _footer(fig, f"Experiment ID: {exp.id}  •  Автоматически сгенерировано")
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def _draw_metrics_page(pdf, flat_metrics, exp_kind=""):
    """One A4 page with a clean table of metrics + explanations + colour-coding."""
    # Build rows: (display_key, label, value_str, explanation, color)
    table_rows = []
    for k, v in flat_metrics.items():
        if not isinstance(v, (int, float)):
            continue
        meta = METRIC_META.get(k, (k.replace("_", " ").title(), "", None))
        label, expl, _ = meta
        color = _metric_traffic_light(k, v)
        fmt_v = f"{v:.4f}" if isinstance(v, float) and abs(v) < 1e6 else str(int(v)) if isinstance(v, int) else str(v)
        table_rows.append((label, fmt_v, expl, color))

    if not table_rows:
        return

    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor(LIGHT_BG)

    title = "Метрики предсказания" if exp_kind == "predict" else "Метрики обучения"
    _title_bar(fig, title, "Сводка ключевых показателей качества модели")

    ax = fig.add_axes([0.04, 0.06, 0.92, 0.85])
    ax.axis("off")

    n = len(table_rows)
    row_h = min(0.115, 0.92 / max(n, 1))

    # Legend
    legend_y = 0.985
    for color, label in [(GREEN, "≥ Хорошо"), (GOLD, "Приемлемо"), (RED, "Ниже нормы")]:
        circ = mpatches.Circle((0, legend_y), 0.008, color=color, transform=ax.transAxes, clip_on=False)
        ax.add_patch(circ)
        ax.text(0.018, legend_y, label, fontsize=7.5, color="#555", va="center", transform=ax.transAxes)
        legend_y  # only show once - reuse x offset
        break  # only one legend strip; handled below

    ax.text(0.0,  0.99, "●", fontsize=10, color=GREEN, va="top", transform=ax.transAxes)
    ax.text(0.03, 0.99, "Хорошо", fontsize=8, color="#555", va="top", transform=ax.transAxes)
    ax.text(0.16, 0.99, "●", fontsize=10, color=GOLD, va="top", transform=ax.transAxes)
    ax.text(0.19, 0.99, "Приемлемо", fontsize=8, color="#555", va="top", transform=ax.transAxes)
    ax.text(0.35, 0.99, "●", fontsize=10, color=RED, va="top", transform=ax.transAxes)
    ax.text(0.38, 0.99, "Требует внимания", fontsize=8, color="#555", va="top", transform=ax.transAxes)
    ax.text(0.62, 0.99, "— (нет порога)", fontsize=8, color="#888", va="top", transform=ax.transAxes)

    y_start = 0.96
    for i, (label, fmt_v, expl, color) in enumerate(table_rows):
        y = y_start - i * row_h
        bg = "#F7F7F7" if i % 2 == 0 else "white"
        rect = mpatches.FancyBboxPatch((-0.01, y - row_h + 0.004), 1.02, row_h - 0.006,
                                        boxstyle="round,pad=0.004", linewidth=0.3,
                                        edgecolor="#E0E0E0", facecolor=bg,
                                        transform=ax.transAxes, clip_on=False)
        ax.add_patch(rect)
        mid = y - row_h * 0.32
        ax.text(0.01, mid, "●", fontsize=11, color=color, va="center", transform=ax.transAxes)
        ax.text(0.045, mid, label, fontsize=9.5, fontweight="bold", color="#222",
                va="center", transform=ax.transAxes)
        ax.text(0.62, mid, fmt_v, fontsize=13, fontweight="bold", color=color,
                va="center", transform=ax.transAxes)
        if expl:
            wrapped = textwrap.fill(expl, width=72)
            ax.text(0.045, y - row_h * 0.74, wrapped, fontsize=7.5, color="#666",
                    va="center", style="italic", transform=ax.transAxes)

    _footer(fig, "Цвет значения показывает соответствие типичным порогам качества для задач бинарной классификации СВ")
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def _draw_graph_page(pdf, graph_type, data):
    """One A4 page per graph with title, description and chart."""
    meta = GRAPH_META.get(graph_type, (graph_type.replace("_", " ").title(), ""))
    title, desc = meta

    fig = plt.figure(figsize=(8.27, 11.69))
    fig.patch.set_facecolor(LIGHT_BG)
    _title_bar(fig, title)

    # Description strip
    ax_desc = fig.add_axes([0.05, 0.88, 0.9, 0.04])
    ax_desc.axis("off")
    ax_desc.text(0.0, 0.5, desc, fontsize=8.5, color="#555", va="center", style="italic")

    ax = fig.add_axes([0.12, 0.12, 0.76, 0.73])
    ax.set_facecolor("white")

    try:
        if graph_type == "roc_curve" and isinstance(data, dict) and "fpr" in data:
            fpr, tpr = data["fpr"], data["tpr"]
            auc_val = data.get("auc")
            ax.plot(fpr, tpr, color=ACCENT, linewidth=2.5,
                    label=f"ROC  AUC = {auc_val:.3f}" if auc_val else "ROC")
            ax.fill_between(fpr, tpr, alpha=0.08, color=ACCENT)
            ax.plot([0, 1], [0, 1], "--", color="#BBBBBB", linewidth=1.2, label="Случайная модель")
            ax.set_xlabel("False Positive Rate", fontsize=11)
            ax.set_ylabel("True Positive Rate", fontsize=11)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.legend(fontsize=10, framealpha=0.9)

        elif graph_type == "pr_curve" and isinstance(data, dict) and "recall" in data:
            rec, prec = data["recall"], data["precision"]
            auc_val = data.get("auc")
            ax.plot(rec, prec, color=GOLD, linewidth=2.5,
                    label=f"PR  AUC = {auc_val:.3f}" if auc_val else "PR")
            ax.fill_between(rec, prec, alpha=0.08, color=GOLD)
            ax.set_xlabel("Recall", fontsize=11)
            ax.set_ylabel("Precision", fontsize=11)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.legend(fontsize=10, framealpha=0.9)

        elif graph_type == "calibration" and isinstance(data, dict) and "probs" in data:
            ax.plot(data["probs"], data["fracs"], "o-", color=GREEN, linewidth=2, markersize=5)
            ax.plot([0, 1], [0, 1], "--", color="#BBBBBB", linewidth=1.2, label="Идеальная калибровка")
            ax.set_xlabel("Предсказанная вероятность", fontsize=11)
            ax.set_ylabel("Реальная частота", fontsize=11)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.legend(fontsize=10)

        elif graph_type == "confusion_matrix" and isinstance(data, dict):
            cm = np.array([
                [data.get("tn", 0), data.get("fp", 0)],
                [data.get("fn", 0), data.get("tp", 0)],
            ], dtype=float)
            im = ax.imshow(cm, cmap="Blues", vmin=0)
            ax.set_xticks([0, 1]); ax.set_yticks([0, 1])
            ax.set_xticklabels(["Pred: Neutral", "Pred: Affected"], fontsize=10)
            ax.set_yticklabels(["True: Neutral", "True: Affected"], fontsize=10)
            for i in range(2):
                for j in range(2):
                    label_cm = [["TN", "FP"], ["FN", "TP"]][i][j]
                    text_color = "black" if cm[i, j] > cm.max() * 0.6 else "black"
                    ax.text(j, i - 0.1, str(int(cm[i, j])), ha="center", va="center",
                            fontsize=16, fontweight="bold", color=text_color)
                    ax.text(j, i + 0.25, f"({label_cm})", ha="center", va="center",
                            fontsize=10, color=text_color, alpha=0.8)
            plt.colorbar(im, ax=ax, fraction=0.04)
            ax.grid(False)

        elif graph_type == "metrics_vs_threshold" and isinstance(data, dict):
            thresholds = data.get("thresholds", [])
            color_map = {"precision": GOLD, "recall": GREEN, "f1": ACCENT}
            for metric_key, color in color_map.items():
                if metric_key in data:
                    ax.plot(thresholds, data[metric_key], color=color, linewidth=2,
                            label=metric_key.capitalize())
            ax.axvline(x=0.5, color=RED, linestyle="--", linewidth=1.2, label="Порог 0.5", alpha=0.7)
            ax.set_xlabel("Порог классификации", fontsize=11)
            ax.set_ylabel("Значение метрики", fontsize=11)
            ax.set_xlim(0, 1); ax.set_ylim(0, 1)
            ax.legend(fontsize=10, framealpha=0.9)

        elif graph_type == "confidence_vs_accuracy" and isinstance(data, dict):
            conf = data.get("confidence", data.get("bins", []))
            acc_vals = data.get("accuracy", [])
            if conf and acc_vals:
                ax.plot(conf, acc_vals, "o-", color=ACCENT, linewidth=2, markersize=5)
                ax.plot([0, 1], [0, 1], "--", color="#BBBBBB", linewidth=1.2, label="Идеальная калибровка")
                ax.set_xlabel("Уверенность модели", fontsize=11)
                ax.set_ylabel("Реальная точность", fontsize=11)
                ax.set_xlim(0, 1); ax.set_ylim(0, 1)
                ax.legend(fontsize=10)

        else:
            # Generic fallback
            colors = ["#2B72FB", "#F2B705", "#27AE60", "#C0392B", "#8E44AD"]
            if isinstance(data, dict):
                for ci, (k, v) in enumerate(data.items()):
                    try:
                        ax.plot(v, label=str(k), color=colors[ci % len(colors)], linewidth=2)
                    except Exception:
                        pass
                ax.legend(fontsize=9)
            else:
                try:
                    ax.plot(data, color=ACCENT, linewidth=2)
                except Exception:
                    ax.text(0.5, 0.5, "Нет данных для отображения", ha="center", va="center",
                            transform=ax.transAxes, color="#999", fontsize=11)

        _set_style(ax)
    except Exception as exc:
        ax.text(0.5, 0.5, f"Ошибка рендеринга:\n{exc}", ha="center", va="center",
                transform=ax.transAxes, color=RED, fontsize=10)

    _footer(fig, f"График: {title}  •  Experiment ID: {data.get('experiment_id', '') if isinstance(data, dict) else ''}")
    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


def _is_float(val):
    try:
        float(val)
        return True
    except Exception:
        return False


# ── Tabular export ────────────────────────────────────────────────────────────

_COL_LABELS = {
    "sv_id":       "SV ID",
    "sv_type":     "Тип СВ",
    "chrom":       "Хромосома",
    "prediction":  "Предсказание",
    "confidence":  "Уверенность (%)",
    "true_label":  "Истинная метка",
    "explanation": "Объяснение",
    "gene_name":   "Ген",
    "name":        "Имя",
}

def _pred_label(val):
    try:
        return "Affected" if int(float(val)) == 1 else "Neutral"
    except Exception:
        return str(val) if val is not None else ""


def _safe_count_predictions(df: pd.DataFrame, label: str) -> int:
    """
    Safely count predictions of a specific label.
    Returns count or "—" if prediction column doesn't exist.
    """
    try:
        if "prediction" not in df.columns:
            return "—"
        return int((df["prediction"].apply(_pred_label) == label).sum())
    except Exception:
        logger.warning(f"Failed to count predictions for {label}, returning '—'")
        return "—"


def _save_predictions_csv(df: pd.DataFrame, out_base: Path):
    csv_path = out_base / "predictions.csv"
    export_df = _prepare_export_df(df)
    export_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    return str(csv_path)


def _prepare_export_df(df: pd.DataFrame) -> pd.DataFrame:
    """Create a clean export copy: human-readable labels, confidence in %."""
    out = df.copy()
    if "prediction" in out.columns:
        out["prediction"] = out["prediction"].apply(_pred_label)
    if "true_label" in out.columns:
        out["true_label"] = out["true_label"].apply(
            lambda v: _pred_label(v) if v not in (None, "", float("nan") if True else None) else ""
        )
    if "confidence" in out.columns:
        out["confidence"] = out["confidence"].apply(
            lambda v: round(float(v) * 100, 2) if _is_float(v) else v
        )
    out.rename(columns=_COL_LABELS, inplace=True)
    return out


def _safe_isna(val) -> bool:
    """Return True if val is NA/NaN/None; never raises for non-scalars."""
    if val is None:
        return True
    try:
        result = pd.isna(val)
        # pd.isna on a scalar returns a bool; on array-like it returns an array
        if isinstance(result, bool):
            return result
        return bool(result)  # single-element arrays
    except (ValueError, TypeError):
        return False


def _to_excel_val(val):
    """Convert a value to a type that openpyxl can serialize."""
    if _safe_isna(val):
        return ""
    if isinstance(val, (pd.Timestamp,)):
        return str(val)
    # numpy scalars are not instances of Python int/float in newer numpy
    try:
        import numpy as np
        if isinstance(val, (np.integer,)):
            return int(val)
        if isinstance(val, (np.floating,)):
            return float(val)
        if isinstance(val, (np.bool_,)):
            return bool(val)
    except ImportError:
        pass
    if not isinstance(val, (str, int, float, bool, type(None))):
        return str(val)
    return val


def _save_predictions_xlsx(df: pd.DataFrame, out_base: Path):
    """Save predictions to XLSX with formatting."""
    try:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed — cannot produce XLSX")

        xlsx_path = out_base / "predictions.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Предсказания"

        cols = list(df.columns)

        # ── Styles ──
        hdr_fill   = PatternFill(patternType="solid", fgColor="1A1A2E")
        hdr_font   = Font(bold=True, color="FFFFFF", size=10)
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side  = Side(style="thin", color="DDDDDD")
        cell_brd   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        fill_even  = PatternFill(patternType="solid", fgColor="F7F7F7")
        fill_odd   = PatternFill(patternType="solid", fgColor="FFFFFF")
        fill_aff   = PatternFill(patternType="solid", fgColor="FFE4E4")
        fill_neu   = PatternFill(patternType="solid", fgColor="E8F5E9")

        # ── Header row ──
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=_COL_LABELS.get(col, col))
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = cell_brd
        ws.row_dimensions[1].height = 28

        # ── Data rows ──
        # Use iterrows() instead of itertuples() so column names with special
        # characters or Python keywords are accessed safely via dict-like row[col].
        pred_ci     = cols.index("prediction") + 1 if "prediction" in cols else None
        expl_ci_idx = cols.index("explanation") if "explanation" in cols else None

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            row_fill = fill_even if ri % 2 == 0 else fill_odd
            for ci, col in enumerate(cols, 1):
                try:
                    raw_val = row.get(col) if hasattr(row, "get") else row[col]

                    # Transform values
                    if col == "prediction":
                        val = _pred_label(raw_val)
                    elif col == "true_label":
                        # Safely handle NaN/None without float("nan") tricks
                        if _safe_isna(raw_val) or str(raw_val).strip() in ("", "nan", "NaN"):
                            val = ""
                        else:
                            val = _pred_label(raw_val)
                    elif col == "confidence":
                        try:
                            val = round(float(raw_val) * 100, 2)
                        except Exception:
                            val = raw_val
                    else:
                        val = raw_val

                    val = _to_excel_val(val)

                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = row_fill
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=(col == "explanation"),
                        horizontal=(
                            "left"   if col == "explanation" else
                            "center" if col in ("prediction", "true_label", "confidence") else
                            "left"
                        ),
                    )
                    cell.border = cell_brd
                except Exception as e:
                    logger.warning(f"Failed to set cell value for row {ri}, column {col}: {str(e)}")
                    cell = ws.cell(row=ri, column=ci, value="[ERROR]")
                    cell.border = cell_brd

            # Colour-code prediction cell
            if pred_ci:
                pred_val = ws.cell(row=ri, column=pred_ci).value
                ws.cell(row=ri, column=pred_ci).fill = fill_aff if pred_val == "Affected" else fill_neu
                ws.cell(row=ri, column=pred_ci).font = Font(
                    bold=True, color="B91C1C" if pred_val == "Affected" else "166534"
                )

            # Auto row height for explanation wrapping
            if expl_ci_idx is not None:
                expl_text = str(row.iloc[expl_ci_idx] if hasattr(row, "iloc") else row[cols[expl_ci_idx]] or "")
                lines = max(1, len(expl_text) // 80 + expl_text.count("\n") + 1)
                ws.row_dimensions[ri].height = max(18, lines * 15)

        # ── Column widths ──
        col_widths = {
            "sv_id":       22, "sv_type": 14, "chrom":  12,
            "prediction":  16, "confidence": 18, "true_label": 16,
            "explanation": 65, "gene_name": 18, "name": 20,
        }
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 16)

        ws.freeze_panes = "A2"

        # ── Summary sheet ──
        ws2 = wb.create_sheet("Сводка")
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 20

        hdr2_font = Font(bold=True, color="FFFFFF", size=10)
        for r, (lbl, val) in enumerate([
            ("Всего предсказаний", len(df)),
            ("Affected (1)", _safe_count_predictions(df, "Affected")),
            ("Neutral (0)",  _safe_count_predictions(df, "Neutral")),
        ], 1):
            c1 = ws2.cell(row=r, column=1, value=lbl)
            c1.font = hdr2_font if r == 1 else Font(bold=True, size=10)
            c1.fill = hdr_fill if r == 1 else (fill_even if r % 2 == 0 else fill_odd)
            c2 = ws2.cell(row=r, column=2, value=val)
            c2.fill = fill_even if r % 2 == 0 else fill_odd

        wb.save(xlsx_path)
        logger.info(f"[EXPORT] XLSX workbook saved to {xlsx_path}")
        return str(xlsx_path)
    except Exception as e:
        logger.error(f"[EXPORT] Failed to create/save XLSX: {str(e)}", exc_info=True)
        raise


def _save_metrics_json(metrics: dict, out_base: Path):
    p = out_base / "metrics.json"
    with open(p, "w", encoding="utf-8") as fh:
        json.dump(metrics, fh, ensure_ascii=False, indent=2)
    return str(p)


def _save_graphs_json(graphs: dict, out_base: Path):
    p = out_base / "graphs.json"
    with open(p, "w", encoding="utf-8") as fh:
        json.dump(graphs, fh, ensure_ascii=False, indent=2)
    return str(p)


def _ensure_out_dir(base: Path):
    base.mkdir(parents=True, exist_ok=True)
    return base


# ── Main entry point ──────────────────────────────────────────────────────────

def export_experiment(experiment_id: int, out_dir: str = None, formats=None):
    """
    Export experiment results.

    formats: list containing any of 'pdf', 'csv', 'xlsx', 'json'
    Returns list of written file paths.

    PDF = professional report: cover + metrics with explanations + graphs.
    CSV / XLSX = predictions table only (full, no pagination).
    JSON = raw metrics + graphs data.
    """
    if formats is None:
        formats = ["pdf"]

    logger.info(f"[EXPORT] Starting export for experiment {experiment_id}, formats={formats}")
    
    exp = Experiment.query.get(experiment_id)
    if exp is None:
        raise RuntimeError(f"Experiment {experiment_id} not found")

    base = (
        Path(out_dir)
        if out_dir
        else Path(__file__).resolve().parent.parent / "exports" / f"experiment_{experiment_id}"
    )
    _ensure_out_dir(base)
    written = []

    # ── Load predictions ──────────────────────────────────────────────────────
    preds_df = None
    if exp.predictions_path and os.path.isfile(exp.predictions_path):
        for sep in ("\t", ","):
            try:
                preds_df = pd.read_csv(exp.predictions_path, sep=sep)
                break
            except Exception:
                pass

    # ── Collect metrics ───────────────────────────────────────────────────────
    metrics = dict(exp.result_data) if exp.result_data else {}
    metric_rows = Metric.query.filter_by(experiment_id=experiment_id).all()
    if metric_rows:
        metrics["metric_rows"] = [m.value for m in metric_rows]

    # ── Collect graphs ────────────────────────────────────────────────────────
    graphs_q = GraphData.query.filter_by(experiment_id=experiment_id).all()
    graphs = {g.graph_type: g.data for g in graphs_q}

    # ── Tabular exports (CSV / XLSX) ──────────────────────────────────────────
    if preds_df is not None:
        if "csv" in formats:
            try:
                written.append(_save_predictions_csv(preds_df, base))
                logger.info(f"[EXPORT] CSV successfully saved for experiment {experiment_id}")
            except Exception as e:
                logger.error(f"[EXPORT] Failed to save CSV for experiment {experiment_id}: {str(e)}", exc_info=True)
        if "xlsx" in formats:
            if not HAS_OPENPYXL:
                logger.warning(f"[EXPORT] openpyxl not installed, cannot export XLSX for experiment {experiment_id}")
            else:
                try:
                    written.append(_save_predictions_xlsx(preds_df, base))
                    logger.info(f"[EXPORT] XLSX successfully saved for experiment {experiment_id}")
                except Exception as e:
                    logger.error(f"[EXPORT] Failed to save XLSX for experiment {experiment_id}: {str(e)}", exc_info=True)

    # ── JSON export ───────────────────────────────────────────────────────────
    if "json" in formats:
        written.append(_save_metrics_json(metrics, base))
        written.append(_save_graphs_json(graphs, base))

    # ── PDF export ────────────────────────────────────────────────────────────
    if "pdf" in formats and HAS_MPL:
        pdf_path = base / "experiment_report.pdf"
        try:
            dataset_name = ""
            try:
                from backend.models import Dataset
                if exp.dataset_id:
                    ds = Dataset.query.get(exp.dataset_id)
                    dataset_name = ds.name if ds else ""
            except Exception:
                pass

            with PdfPages(pdf_path) as pdf:

                # 1. Cover page
                _draw_cover_page(pdf, exp, dataset_name)

                # 2. Metrics page — flatten nested dicts (e.g. test_metrics)
                flat_metrics = {}
                for k, v in metrics.items():
                    if k == "metric_rows":
                        continue
                    if isinstance(v, dict):
                        flat_metrics.update({sk: sv for sk, sv in v.items() if isinstance(sv, (int, float))})
                    elif isinstance(v, (int, float)):
                        flat_metrics[k] = v

                if flat_metrics:
                    _draw_metrics_page(pdf, flat_metrics, exp.kind)

                # 3. Graph pages in a consistent order
                GRAPH_ORDER = [
                    "roc_curve", "pr_curve", "confusion_matrix",
                    "calibration",
                    "metrics_vs_threshold", "confidence_vs_accuracy",
                ]
                ordered = [(k, graphs[k]) for k in GRAPH_ORDER if k in graphs]
                ordered += [(k, v) for k, v in graphs.items() if k not in GRAPH_ORDER]

                for gtype, gdata in ordered:
                    try:
                        _draw_graph_page(pdf, gtype, gdata)
                    except Exception:
                        continue

            written.append(str(pdf_path))
        except Exception as e:
            logger.error(f"[EXPORT] Failed to save PDF for experiment {experiment_id}: {str(e)}", exc_info=True)

    logger.info(f"[EXPORT] Export completed for experiment {experiment_id}: {len(written)} files written")
    return written
